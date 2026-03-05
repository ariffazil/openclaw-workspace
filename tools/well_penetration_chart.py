#!/usr/bin/env python3
"""
Well Penetration Chart Generator
=================================
Converts messy biostratigraphic pick data into a clean, visually-attractive Excel matrix.

Dependencies:
    pip install pandas openpyxl

Usage:
    python well_penetration_chart.py                        # uses built-in sample data
    python well_penetration_chart.py picks.tsv              # reads TSV file
    python well_penetration_chart.py picks.tsv output.xlsx  # custom output path

Input TSV columns: Region, Well, PickName, Depth_m

Output Excel:
    Sheet 1 "Well Penetration Matrix" — the deliverable chart
    Sheet 2 "Cleaned Picks"           — audit trail of what was parsed
"""

import io
import re
import sys

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# ZONE EXTRACTION
# =============================================================================

def extract_canonical_zone(pick_name: str) -> str | None:
    """
    Extract canonical NN zone from a messy PickName string.

    Strategy:
    - Find all NN\\d+ occurrences via regex.
    - Take the FIRST match (= youngest / shallowest in a range like "NN10-NN9").
    - Strip subzone letters: NN9A → NN9, NN11E → NN11.
    - Return None if no NN zone found (row will be dropped with a warning).

    Examples:
        'SB_Biostrat_NN11'               → 'NN11'
        'NN10 - NN9'                     → 'NN10'
        '?NN8' / '_NN8'                  → 'NN8'
        'NN9A' / 'SB_Biostrat_NN9A'      → 'NN9'
        'IVD_NN11'                       → 'NN11'
        'Base Stage IVD_NN10_11_...'     → 'NN10'
        'Kamunsu Top_NN11A_11B_...'      → 'NN11'
        'NN9/?NN8'                       → 'NN9'
        'SB_Biostrat_NN9A-NN8'           → 'NN9'
    """
    hits = re.findall(r'NN(\d+)', str(pick_name), re.IGNORECASE)
    if not hits:
        return None
    return f"NN{int(hits[0])}"


def sort_key(zone: str) -> int:
    """Higher NN number = younger age = appears at TOP of chart."""
    m = re.match(r'NN(\d+)', zone, re.IGNORECASE)
    return int(m.group(1)) if m else 0


# =============================================================================
# DATA PROCESSING
# =============================================================================

def process(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean raw picks into a tidy table.
    Returns df_tops: one row per (Well, canonical_zone), depth = zone top (min).
    """
    df = df.copy()
    df["canonical_zone"] = df["PickName"].apply(extract_canonical_zone)

    unmapped = df[df["canonical_zone"].isna()]["PickName"].unique()
    for u in unmapped:
        print(f"  [WARN] No NN zone found in: '{u}'", file=sys.stderr)
    df = df.dropna(subset=["canonical_zone"])

    # Per Well × Zone: shallowest = top of zone (first encountered while drilling)
    df_tops = (
        df.groupby(["Well", "canonical_zone"], as_index=False)["Depth_m"]
        .min()
        .rename(columns={"Depth_m": "Top_Depth_m"})
    )
    return df_tops


def build_pivot(df_tops: pd.DataFrame) -> pd.DataFrame:
    """Pivot into (Zone rows) × (Well columns), youngest zone at top."""
    pivot = df_tops.pivot_table(
        index="canonical_zone",
        columns="Well",
        values="Top_Depth_m",
        aggfunc="min",
    )
    sorted_zones = sorted(pivot.index, key=sort_key, reverse=True)
    pivot = pivot.loc[sorted_zones].sort_index(axis=1)
    return pivot


# =============================================================================
# COLOR UTILITIES
# =============================================================================

# Depth gradient: shallow = light mint, deep = dark emerald
_SHALLOW_RGB = (209, 250, 229)   # #D1FAE5
_DEEP_RGB    = (6,   78,  59)    # #064E3B


def depth_to_hex(depth: float, d_min: float, d_max: float) -> tuple[str, bool]:
    """
    Map a depth value to a hex colour string using a shallow→deep green gradient.
    Returns (hex_color, use_white_text).
    """
    t = (depth - d_min) / (d_max - d_min) if d_max != d_min else 0.5
    t = max(0.0, min(1.0, t))
    r = int(_SHALLOW_RGB[0] + t * (_DEEP_RGB[0] - _SHALLOW_RGB[0]))
    g = int(_SHALLOW_RGB[1] + t * (_DEEP_RGB[1] - _SHALLOW_RGB[1]))
    b = int(_SHALLOW_RGB[2] + t * (_DEEP_RGB[2] - _SHALLOW_RGB[2]))
    return f"{r:02X}{g:02X}{b:02X}", t > 0.50


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, color="111111", size=9, italic=False, name="Calibri") -> Font:
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)


def _border(color="CCCCCC", style="thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border(color="888888") -> Border:
    thick = Side(style="medium", color=color)
    thin  = Side(style="thin",   color="DDDDDD")
    return Border(left=thick, right=thick, top=thin, bottom=thin)


# =============================================================================
# EXCEL WRITER
# =============================================================================

def write_excel(pivot: pd.DataFrame, output_path: str) -> None:
    wells = list(pivot.columns)
    zones = list(pivot.index)

    # Global depth range for gradient
    flat = [v for v in pivot.values.flatten() if not pd.isna(v)]
    d_min, d_max = (min(flat), max(flat)) if flat else (0, 1)

    # TD zone per well = zone with the maximum depth value
    td_zone: dict[str, str] = {}
    for well in wells:
        col = pivot[well].dropna()
        if not col.empty:
            td_zone[well] = col.idxmax()

    wb = openpyxl.Workbook()

    # =========================================================== SHEET 1
    ws = wb.active
    ws.title = "Well Penetration Matrix"

    # Print setup — landscape, fit to one page wide
    ws.page_setup.orientation       = "landscape"
    ws.page_setup.fitToPage         = True
    ws.page_setup.fitToWidth        = 1
    ws.page_setup.fitToHeight       = 0
    ws.page_margins.left            = 0.4
    ws.page_margins.right           = 0.4
    ws.page_margins.top             = 0.5
    ws.page_margins.bottom          = 0.5
    ws.print_title_rows             = "1:3"   # repeat header rows when printed

    NCOLS      = len(wells) + 1
    HDR_ROW    = 3   # row with well names
    DATA_START = 4   # first data row

    # ---- Row 1: Title bar -----------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    t = ws.cell(row=1, column=1,
                value="WELL PENETRATION CHART  ·  SB Region  ·  NN Nannofossil Biozones")
    t.font      = _font(bold=True, color="FFFFFF", size=13)
    t.fill      = _fill("1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---- Row 2: Subtitle bar --------------------------------------------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    s = ws.cell(row=2, column=1,
                value=(
                    f"Depths in metres (m)  ·  Values = top of zone  ·  "
                    f"Colour intensity = drilling depth ({int(d_min)} m shallow → {int(d_max)} m deep)  ·  "
                    f"Red cell ▼ = Total Depth (TD)  ·  White cell = not penetrated"
                ))
    s.font      = _font(italic=True, color="AAAAAA", size=8)
    s.fill      = _fill("111D2E")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 13

    # ---- Row 3: Column headers (well names, rotated 90°) ----------------
    # "BIOZONE" corner cell
    corner = ws.cell(row=HDR_ROW, column=1, value="BIOZONE")
    corner.fill      = _fill("253F6A")
    corner.font      = _font(bold=True, color="FFFFFF", size=9)
    corner.alignment = Alignment(horizontal="center", vertical="center")
    corner.border    = _border("1F3864", "medium")
    ws.row_dimensions[HDR_ROW].height = 90   # tall row to fit rotated text

    for c_i, well in enumerate(wells, start=2):
        h = ws.cell(row=HDR_ROW, column=c_i, value=well)
        h.fill      = _fill("253F6A")
        h.font      = _font(bold=True, color="FFFFFF", size=9)
        # textRotation=90 → text reads bottom-to-top (standard for column charts)
        h.alignment = Alignment(
            horizontal="center", vertical="bottom",
            textRotation=90, wrap_text=False
        )
        h.border = _border("1F3864", "medium")

    # ---- Data rows -------------------------------------------------------
    for r_i, zone in enumerate(zones, start=DATA_START):
        ws.row_dimensions[r_i].height = 22

        # Zone label
        zc = ws.cell(row=r_i, column=1, value=zone)
        zc.fill      = _fill("2E5FA3")
        zc.font      = _font(bold=True, color="FFFFFF", size=10)
        zc.alignment = Alignment(horizontal="center", vertical="center")
        zc.border    = _thick_border("1A3D7A")

        for c_i, well in enumerate(wells, start=2):
            depth_val = pivot.loc[zone, well]
            cell = ws.cell(row=r_i, column=c_i)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if pd.isna(depth_val):
                # Not penetrated — white, barely-visible border
                cell.fill   = _fill("FFFFFF")
                cell.value  = ""
                cell.border = _border("E8E8E8", "hair")
                cell.font   = _font(color="DDDDDD", size=8)

            else:
                depth  = float(depth_val)
                is_td  = (td_zone.get(well) == zone)

                if is_td:
                    # Total Depth — strong red, bold
                    cell.fill      = _fill("C62828")
                    cell.font      = _font(bold=True, color="FFFFFF", size=9)
                    cell.value     = f"{depth:,.0f}\n▼"
                    cell.border    = Border(
                        left   = Side(style="medium", color="8B0000"),
                        right  = Side(style="medium", color="8B0000"),
                        top    = Side(style="medium", color="8B0000"),
                        bottom = Side(style="medium", color="8B0000"),
                    )
                    ws.row_dimensions[r_i].height = 30

                else:
                    hex_c, white_text = depth_to_hex(depth, d_min, d_max)
                    cell.fill   = _fill(hex_c)
                    cell.font   = _font(
                        size=9,
                        color="FFFFFF" if white_text else "0D3320"
                    )
                    cell.value  = f"{depth:,.0f}"
                    cell.border = _border("B0C4B0", "thin")

    # ---- Thick bottom border on last data row ---------------------------
    last_row = DATA_START + len(zones) - 1
    for c_i in range(1, NCOLS + 1):
        cell = ws.cell(row=last_row, column=c_i)
        b = cell.border
        cell.border = Border(
            left=b.left, right=b.right, top=b.top,
            bottom=Side(style="medium", color="444444")
        )

    # ---- Column widths --------------------------------------------------
    ws.column_dimensions["A"].width = 9      # biozone label
    for c_i in range(2, NCOLS + 1):
        ws.column_dimensions[get_column_letter(c_i)].width = 8  # narrow — rotated text fits

    ws.freeze_panes = "B4"   # freeze zone column + 2 header rows

    # ---- Legend (gradient bar + TD swatch) ------------------------------
    LEG_ROW = DATA_START + len(zones) + 2
    ws.row_dimensions[LEG_ROW].height = 16

    lbl = ws.cell(row=LEG_ROW, column=1, value="Depth →")
    lbl.font      = _font(bold=True, color="555555", size=8)
    lbl.alignment = Alignment(horizontal="right", vertical="center")

    N_GRAD = min(9, NCOLS - 2)   # gradient swatches — don't overflow columns
    for i in range(N_GRAD):
        t_val = i / max(N_GRAD - 1, 1)
        depth_at_t = d_min + t_val * (d_max - d_min)
        hex_c, white_text = depth_to_hex(depth_at_t, d_min, d_max)
        gc = ws.cell(row=LEG_ROW, column=2 + i)
        gc.fill      = _fill(hex_c)
        gc.alignment = Alignment(horizontal="center", vertical="center")
        gc.border    = _border("BBBBBB", "thin")
        if i == 0:
            gc.value = f"{int(d_min)} m"
            gc.font  = _font(size=7, color="0D3320")
        elif i == N_GRAD - 1:
            gc.value = f"{int(d_max)} m"
            gc.font  = _font(size=7, color="FFFFFF")

    # TD swatch
    td_col = 2 + N_GRAD + 1
    td_swatch = ws.cell(row=LEG_ROW, column=td_col, value="▼ TD")
    td_swatch.fill      = _fill("C62828")
    td_swatch.font      = _font(bold=True, color="FFFFFF", size=8)
    td_swatch.alignment = Alignment(horizontal="center", vertical="center")
    td_swatch.border    = _border("8B0000", "thin")

    empty_swatch = ws.cell(row=LEG_ROW, column=td_col + 1, value="not penetrated")
    empty_swatch.fill      = _fill("FFFFFF")
    empty_swatch.font      = _font(color="AAAAAA", size=8)
    empty_swatch.alignment = Alignment(horizontal="center", vertical="center")
    empty_swatch.border    = _border("E8E8E8", "hair")

    # =========================================================== SHEET 2
    ws2 = wb.create_sheet("Cleaned Picks")
    headers2 = ["Well", "Canonical Zone", "Zone Top (m)", "Is TD?"]
    for c, h in enumerate(headers2, 1):
        hc = ws2.cell(row=1, column=c, value=h)
        hc.fill      = _fill("1F3864")
        hc.font      = _font(bold=True, color="FFFFFF", size=10)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border    = _border("1F3864", "thin")

    r = 2
    for well in wells:
        for zone in zones:
            val = pivot.loc[zone, well]
            if not pd.isna(val):
                is_td = td_zone.get(well) == zone
                ws2.cell(row=r, column=1, value=well)
                ws2.cell(row=r, column=2, value=zone)
                ws2.cell(row=r, column=3, value=round(float(val), 1))
                ws2.cell(row=r, column=4, value="YES" if is_td else "")
                if is_td:
                    for c in range(1, 5):
                        ws2.cell(row=r, column=c).fill = _fill("FECDD3")
                r += 1

    for i, w in enumerate([22, 16, 16, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"[DONE] Saved → {output_path}", file=sys.stderr)


# =============================================================================
# BUILT-IN SAMPLE DATA
# =============================================================================

SAMPLE = """\
Region\tWell\tPickName\tDepth_m
SB\tBRUNEI BAY-1R1\tSB_Biostrat_NN11\t329.84024
SB\tBRUNEI BAY-1R1\tNN9\t1000
SB\tBRUNEI BAY-1R1\tSB_Biostrat_NN9\t1000
SB\tDELUAR-1\tNN5\t1287.611
SB\tDELUAR-1\tNN9\t579.58154
SB\tDUDAR-1\tNN11\t1370
SB\tDUDAR-1\tNN5\t2368
SB\tDUDAR-1\tNN9\t1737
SB\tJANGAS-2S1\tNN10 - NN9\t2900
SB\tJANGAS-2S1\tSB_Biostrat_NN10_NN9\t2900
SB\tJANGAS-2S1\tNN11\t2600
SB\tJANGAS-2S1\tSB_Biostrat_NN11\t2600
SB\tJANGAS-2S1\tNN15 - NN12\t2550
SB\tJANGAS-2S1\tSB_Biostrat_NN15_NN12\t2550
SB\tJANGAS-2S1\tNN10 - NN9\t2900
SB\tJANGAS-2S1\tSB_Biostrat_NN10_NN9\t2900
SB\tKARUKAN-1\tNN9\t1112
SB\tMADALON-1\tNN11\t967
SB\tMADALON-1\tNN15\t427.00003
SB\tPISAGAN-1A\tNN10\t2223.48
SB\tPISAGAN-1A\tNN10\t2223.48
SB\tPISAGAN-1A\tBase Stage IVD_NN10_11_WellReports\t2484
SB\tPISAGAN-1A\tNN10\t2484
SB\tPISAGAN-1A\tSB_Biostrat_NN10_NN9\t2484
SB\tPISAGAN-1A\tNN11\t2017.07
SB\tPISAGAN-1A\tNN11\t2017.776
SB\tPISAGAN-1A\tNN11E\t2018
SB\tPISAGAN-1A\tBase Stage IVE_NN11_WellReports\t2049.309
SB\tPISAGAN-1A\tKamunsu Top_NN11A_11B_WellReports\t2294.9414
SB\tPISAGAN-1A\tNN9/?NN8\t3169.68
SB\tPISAGAN-1A\tSB_Biostrat_NN9A-NN8\t3170
SB\tPISAGAN-1A\t?NN8\t3306.87
SB\tPISAGAN-1A\tNN8\t3307
SB\tPISAGAN-1A\tSB_Biostrat_NN8\t3307
SB\tPISAGAN-1A\t_NN8\t3307.08
SB\tPISAGAN-1A\tSB_Biostrat_NN10_NN9\t2484
SB\tPISAGAN-1A\tNN9A\t2895.52
SB\tPISAGAN-1A\tNN9\t2895.6
SB\tPISAGAN-1A\tNN9A\t2896
SB\tPISAGAN-1A\tSB_Biostrat_NN9A\t2896
SB\tPISAGAN-1A\tNN9/?NN8\t3169.68
SB\tPISAGAN-1A\tSB_Biostrat_NN9A-NN8\t3170
SB\tROTAN-1\tNN10\t1993.9177
SB\tROTAN-1\tIVD_NN11\t1836.3722
SB\tSOLISIP-1\tNN11\t877
SB\tSOLISIP-1\tNN7\t1899.4757
SB\tSUGUT-1\tNN10\t555
SB\tSUGUT-1\tNN7\t582
SB\tTEKUYONG-1\tNN11\t1390
"""


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    if len(sys.argv) >= 2:
        src = sys.argv[1]
        sep = "\t" if src.lower().endswith((".tsv", ".txt")) else ","
        df  = pd.read_csv(src, sep=sep)
        print(f"[INFO] Loaded {len(df)} rows from '{src}'", file=sys.stderr)
    else:
        df = pd.read_csv(io.StringIO(SAMPLE), sep="\t")
        print(f"[INFO] Using built-in sample data ({len(df)} rows)", file=sys.stderr)

    out = sys.argv[2] if len(sys.argv) >= 3 else "well_penetration_chart.xlsx"

    df_tops = process(df)
    pivot   = build_pivot(df_tops)

    print(f"[INFO] Matrix: {pivot.shape[0]} zones × {pivot.shape[1]} wells", file=sys.stderr)
    write_excel(pivot, out)
